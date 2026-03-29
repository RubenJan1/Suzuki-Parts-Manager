# ================================================================
# engine_tradelist.py
# Gebaseerd 1-op-1 op jouw tradelist_maker_PERFECT_2025_ECHT_KLAAR
# Logica is NIET gewijzigd – alleen ingepakt als engine voor de app
# ================================================================

import pandas as pd
import re
import html
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from pathlib import Path
from utils.paths import output_root

class TradelistEngine:
    def __init__(self, app_state, wc_csv_path=None):
        if app_state.wc_df is None or app_state.wc_df.empty:
            raise RuntimeError("WC-export niet geladen")

        self.app_state = app_state
        self.df = app_state.wc_df.copy()
        self.wc_csv_path = wc_csv_path

        RE_K_LOC = re.compile(r"^K\s*\d+", re.I)



    # --------------------------------------------------------------
    # HOOFD RUN – identiek gedrag als jouw script
    # --------------------------------------------------------------
    def run(self, output_dir=str(output_root() / "tradelist")):
        df = self.df.copy()

        df.columns = df.columns.astype(str).str.strip()
        # ----------------------------------------------------------
        # 1. Instellingen
        # ----------------------------------------------------------
        current_date = datetime.now()
        datum_str = current_date.strftime('%m-%y')

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        OUTPUT_ONSZELF = output_dir / f"TRADELIST-ONSZELF-{datum_str}.xlsx"
        OUTPUT_CMS = output_dir / f"CMS-TRADELIST-{datum_str}.xlsx"
        OUTPUT_GOPARTS = output_dir / f"GOPARTS-TRADELIST-{datum_str}.xlsx"
        LOG_XLSX = output_dir / f"tradelist_LOG_VERWIJDERD_{current_date.strftime('%Y%m%d_%H%M')}.xlsx"

        KORTING_CMS = 0.86
        KORTING_GOPARTS = 0.81

        log_lines = []
        verwijderd = []

        def verwijder(titel, reden):
            lijn = f"VERWIJDERD → {titel} | {reden}"
            verwijderd.append(lijn)
            log_lines.append(lijn)

        log_lines.append(
            f"=== TRADELIST PERFECT 2025 – ALLEEN JOUW REGELS === {current_date.strftime('%d-%m-%Y %H:%M')}"
        )

        RE_K_LOC = re.compile(r"^K\s*\d+", re.I)

        def is_suzuki_categorie(cat: str) -> bool:
            c = str(cat or "").upper()
            # Alles onder "Verschillende merken" is NIET Suzuki
            if "VERSCHILLENDE MERKEN" in c:
                return False
            # Jouw Suzuki-boom zit onder "Originele onderdelen"
            if "ORIGINELE ONDERDELEN" in c:
                return True
            # fallback (optioneel): herken Suzuki modelfamilies
            suz_tokens = ["GSX", "GSX-R", "GT", "GS SERIES", "DR", "SP", "RM", "RG", "TS", "VS", "VZ", "VX", "VL", "T SERIES"]
            return any(t in c for t in suz_tokens)
        
        RE_HTML_TAGS = re.compile(r"<[^>]+>")


        def clean_text(x: str) -> str:
            """
            Maak WooCommerce-tekst schoon:
            - HTML tags weg
            - HTML entities decoden
            - \n / \r normaliseren
            - lege regels verwijderen
            - whitespace normaliseren
            """
            if x is None:
                return ""

            s = str(x)

            # 1) HTML entities (&amp; &nbsp; etc)
            s = html.unescape(s)

            # 2) HTML tags (<br>, <p>, etc)
            s = RE_HTML_TAGS.sub(" ", s)

            # 3) Newlines en carriage returns -> spatie
            s = s.replace("\r", " ").replace("\n", " ")
            s = s.replace("\\r", " ").replace("\\n", " ")

            # 4) Non-breaking spaces
            s = s.replace("\u00a0", " ")

            # 5) Meerdere spaties / tabs -> 1 spatie
            s = re.sub(r"[ \t]+", " ", s)

            # 6) Trim
            return s.strip()
        
        def is_derbi_row(row) -> bool:
            cat = str(row.get("Cat", "") or "").upper()
            merk_csv = str(row.get("Merk_CSV", "") or "").upper()
            full = f"{row.get('Locatie','')} {row.get('KB','')} {row.get('Cat','')} {row.get('Merk_CSV','')}".upper()
            return ("DERBI" in cat) or ("DERBI" in merk_csv) or (re.search(r"\bDERBI\b", full) is not None)



        # ----------------------------------------------------------
        # 2. Kolommen hernoemen (IDENTIEK)
        # ----------------------------------------------------------
        df = df.fillna('').rename(columns={
            'Naam': 'Title',
            'Reguliere prijs': 'Prijs',
            'Voorraad': 'Stock',
            'Korte beschrijving': 'KB',
            'Beschrijving': 'Locatie',
            'Categorieën': 'Cat',
            'Merken': 'Merk_CSV'
        })

        if 'Merk_CSV' not in df.columns:
            df['Merk_CSV'] = ''
            
        df = df[['Title', 'Prijs', 'Stock', 'KB', 'Locatie', 'Cat', 'Merk_CSV']].copy()
        # HTML/entiteiten opschonen in tekstvelden
        # HTML / tekst opschonen
        for col in ["Title", "KB", "Locatie", "Cat", "Merk_CSV"]:
            df[col] = df[col].apply(clean_text)



        df['Prijs'] = pd.to_numeric(
            df['Prijs'].astype(str).str.replace(r'[^\d,.]', '', regex=True).str.replace(',', '.'),
            errors='coerce'
        ).fillna(0)

        df['Stock'] = pd.to_numeric(df['Stock'], errors='coerce').fillna(0).astype(int)

        totaal = len(df)
        df = df[(df['Prijs'] > 0) & (df['Stock'] >= 1)].copy()
        if totaal - len(df):
            verwijder('(meerdere)', f"Geen prijs of voorraad → {totaal - len(df)} verwijderd")

        # ----------------------------------------------------------
        # 4. Troep filter (UPGRADED: blacklist + partnumber whitelist)
        # ----------------------------------------------------------

        RE_PATTERN_IMI = re.compile(
            r"\b(pattern|patern|patten|imi|imitation|imitatie|namaak|fake|replica|aftermarket|copy|repro)\b",
            re.I
        )


        RE_DOC_KEYWORDS = re.compile(
            r"\b("
            r"manual|handleiding|handbuch|fahrerhandbuch|werkplaatshandboek|werkplaats|workshop|service\s*manual|"
            r"parts?\s*catalog(ue)?|partscatalogue|catalogue\s*copy|"
            r"wiring\s*diagrams?|schema|schaltplan|"
            r"technical\s*bulletin|technical\s*information|technische|introduktie|introductie|"
            r"brochure|leaflet|"
            r"uso\s*e\s*manutenzione|wartungsanleitung|"
            r"lecture\s*handbook|service\s*data|setup\s*manual|"
            r"key\s*number|keynumber"
            r")\b",
            re.I
        )

        # Title moet een partnumber zijn (jouw business rule)
        RE_ALLOWED_CHARS = re.compile(r"^[0-9A-Z\-]+$", re.I)
        RE_DERBI_NODASH = re.compile(r"^(?=.{8,12}$)[0-9A-Z]+$", re.I)

        # Bekende formats (niet té strict, maar wel effectief)
        # Hyphenated nummers: bv 09440-19027, 4L0-14610-00, 990D0-06G00-030
        RE_PART_HYPHEN = re.compile(
            r"^(?=.{5,35}$)"                 # lengte
            r"[0-9A-Z]{1,8}"                 # blok 1
            r"(?:-[0-9A-Z]{1,8}){1,5}$",     # 1 t/m 5 extra blokken
            re.I
        )

        # Non-hyphen codes: bv 130G0660, 670B2014, 00G02209252, 00004184500
        RE_PART_NODASH = re.compile(
            r"^(?=.{6,20}$)"          # lengte
            r"(?=(?:.*\d){6,})"       # minstens 6 cijfers totaal
            r"\d[0-9A-Z]+$",          # moet met cijfer beginnen
            re.I
        )




        def normalize_title(x: str) -> str:
            t = str(x or "").strip().upper()
            t = re.sub(r"\s+", " ", t)
            t = re.sub(r"\b(ORANGE|BLACK|WHITE|RED|BLUE|GREEN|YELLOW|SILVER|GOLD)\b$", "", t).strip()
            return t

        def is_partnumber(title: str) -> bool:
            t = normalize_title(title).replace(" ", "")  # geen spaties in codes
            if not t:
                return False

            # Alleen A-Z/0-9/- toegestaan
            if not RE_ALLOWED_CHARS.match(t):
                return False

            # jouw harde regels
            if t.endswith("-P") or t.endswith("-KEY"):
                return False

            # Hyphen of no-dash formaten accepteren
            if RE_PART_HYPHEN.match(t):
                return True
            if RE_PART_NODASH.match(t):
                return True

            return False


        
        # NOTE: keep these as word/phrase matches to avoid false positives like "thrust" (contains "rust")
        # or "defector" (contains "defect").
        RE_CONDITION = re.compile(
            r"\b("
            r"(2nd|second)\s*hand|"
            r"used|briefly\s*used|"
            r"(as\s+)?good\s+as\s+new|"                       # (as) good as new
            r"shortly\s*(mounted|installed|fitted)|"          # shortly mounted
            r"briefly\s*(mounted|installed|fitted)|"          # briefly mounted
            r"(test|trial)\s*(mounted|installed|fitted|fit)|" # test/trial fitted
            r"complete\s+with\b.{0,40}\bmounted|"             # complete with ... mounted
            r"bearings?\s+mounted|"                           # bearings mounted
            r"few\s+minor\s+marks?|"                          # few minor marks
            r"few\s+minor\s+scuffs?|"                         # few minor scuffs
            r"minor\s+chrome\s+pitting|"                      # minor chrome pitting
            r"nos\s+with\s+minor\s+pitting|"                  # NOS WITH MINOR PITTING
            r"used\s*but\s*(intact|intakt|checked|working|very\s*nice|like\s*new|100\s*%|100%)|"
            r"used\s*bur|"                                    # historical typo case
            r"like\s*new|"
            r"refurbished|reconditioned|overhauled|repaired|restored|"  # common b2b wording
            r"\brust(y|ed)?\b|\bcorrosion\b|\boxidation\b|\bpitted\b|"  # word-boundary safety
            r"\bdefect(ive)?\b|\bfaulty\b|\bnot\s+working\b|"           # word-boundary safety
            r"scuff(ed|ing)?|scratch(ed|es|ing)?|dent(ed)?|"
            r"damage(d)?|"
            r"needs?\s*new\s*paint(job)?|needs?\s*paint(job)?|"
            r"minor\s*paint|"
            r"paint\b.{0,30}\b(damage|damaged|marks?|scratch|scratched|scuff|scuffed|dent|chip)\b|"
            r"some\s*marks?\s*on\s*the\s*paint|"
            r"few\s*paint\s*marks"
            r")\b",
            re.I
        )

        RE_TITLE_EXCLUDE = re.compile(
            r"\b("
            r"parts?\s+catalo(?:g|q)ue|parts?\s+catalog|"  # catalogue + cataloque + catalog
            r"spark\s*plug|sparkplug|"
            r"tuning|"
            r"workshop\s*manual|service\s*manual|manual|"
            r"wiring\s*diagrams?|wiring\s*diagram|schematic|schema"
            r")\b",
            re.I
        )

        def is_troep(row):
            title_raw = row.get("Title", "")
            t = str(title_raw).lower().strip()
            kb = str(row.get("KB", "")).lower()
            loc = str(row.get("Locatie", "")).lower()
            cat = str(row.get("Cat", "")).lower()
            full = f"{t} {kb} {loc} {cat}"

            loc_raw = str(row.get("Locatie", "")).strip()
            cat_raw = str(row.get("Cat", ""))

            # 0) Niet-onderdelen op Title
            if RE_TITLE_EXCLUDE.search(row.get("Title", "")):
                verwijder(row["Title"], "Niet-onderdeel (catalogue/sparkplug/tuning/manual/diagram) in Title")
                return True
            
            if RE_K_LOC.match(loc_raw) and is_suzuki_categorie(cat_raw):
                verwijder(row["Title"], f"Suzuki met K-locatie ({loc_raw}) - locatie fout/oud systeem")
                return True

            # 1) Pattern/IMI (incl. typos)
            if RE_PATTERN_IMI.search(full):
                verwijder(row["Title"], "Pattern/IMI gevonden (incl. typo)")
                return True

            # 2) Documentatie/keys/bulletins/etc (multilingual)
            if RE_DOC_KEYWORDS.search(full):
                verwijder(row["Title"], "Documentatie/Key/Bulletin/etc gevonden")
                return True
            
            # 2b) Conditie / schade / used / 2nd hand
            # NOS moet NIET verwijderd worden, maar NOS + "used/scuffed/damage" is alsnog niet wenselijk.
            if RE_CONDITION.search(full):
                verwijder(row["Title"], "Conditie-aanduiding (used/damaged/scuffed/etc) gevonden")
                return True
            # 3) Extra specifieke suffix regels
            tt = normalize_title(row["Title"]).replace(" ", "")
            if tt.endswith("-P"):
                verwijder(row["Title"], "Eindigt op -P")
                return True
            if tt.endswith("-KEY"):
                verwijder(row["Title"], "Eindigt op -KEY")
                return True

            # 4) Hard whitelist: Title moet partnumber-formaat hebben
            if not is_partnumber(row["Title"]):
                verwijder(row["Title"], "Title is geen onderdeelnummer (whitelist)")
                return True

            return False

        df = df[~df.apply(is_troep, axis=1)].copy()


        # ----------------------------------------------------------
        # 5. Merk bepalen (ONGEWJIZIGD)
        # ----------------------------------------------------------
        def bepaal_merk(row):
            merk_csv = str(row.get("Merk_CSV") or "").strip()
            if merk_csv:
                return merk_csv.title()

            loc = str(row.get("Locatie") or "").strip()
            kb  = str(row.get("KB") or "").strip()
            cat = str(row.get("Cat") or "").strip()

            full_u = f"{loc} {kb} {cat}".upper()
            cat_u  = cat.upper()
            loc_u  = loc.upper()

            # 1) Derbi alleen als het woord echt voorkomt (D0/D1 = DOOS!)
            if re.search(r"\bDERBI\b", full_u):
                return "Derbi"

            # 2) Categorieën zijn leidend voor niet-Suzuki
            #    (bij jou staan andere merken meestal onder "Verschillende merken > <Merk>")
            if "VERSCHILLENDE MERKEN" in cat_u:
                if "HONDA" in cat_u: return "Honda"
                if "KAWASAKI" in cat_u: return "Kawasaki"
                if "YAMAHA" in cat_u: return "Yamaha"
                if "DERBI" in cat_u: return "Derbi"

            # 3) Expliciete merknaam in tekst (loc/kb/cat)
            if "HONDA" in full_u: return "Honda"
            if "KAWASAKI" in full_u: return "Kawasaki"
            if "YAMAHA" in full_u: return "Yamaha"

            # 4) Locatie-prefix alleen gebruiken als cat "neutraal" is (dus geen Suzuki modelboom)
            #    Hiermee voorkom je dat Suzuki in K-dozen ineens "Kawasaki" wordt.
            is_suzuki_cat = "ORIGINELE ONDERDELEN" in cat_u and "VERSCHILLENDE MERKEN" not in cat_u
            if not is_suzuki_cat:
                m = re.match(r"^([HKY])\s*\d+", loc_u)
                if m:
                    return {"H": "Honda", "K": "Kawasaki", "Y": "Yamaha"}[m.group(1)]

            return "Suzuki"




        # ----------------------------------------------------------
        # 6. Categorie opschonen
        # ----------------------------------------------------------
        def schoon_categorieen(txt):
            if not txt:
                return ''
            cats = [c.strip() for c in txt.split(',') if c.strip()]
            seen = set()
            uniq = [c for c in cats if not (c in seen or seen.add(c))]
            return ' > '.join(uniq)

        df['Cat_Schoon'] = df['Cat'].apply(schoon_categorieen)

        # ----------------------------------------------------------
        # 7. Prijzen
        # ----------------------------------------------------------
        def excel_safe(text):
            if isinstance(text, str) and text and text[0] in ('=', '+', '-', '@'):
                return "'" + text
            return text

        df['Prijs_Jouw'] = df['Prijs'].round(2)
        df['Prijs_CMS'] = (df['Prijs'] * KORTING_CMS).round(2)
        df['Prijs_GoParts'] = (df['Prijs'] * KORTING_GOPARTS).round(2)

        # ----------------------------------------------------------
        # 8. Output dataframes
        # ----------------------------------------------------------
        if 'Merk' not in df.columns:
            if 'Merk_CSV' not in df.columns:
                df['Merk_CSV'] = ''
            df['Merk'] = df.apply(bepaal_merk, axis=1)
        df_onszelf = df[['Title', 'Merk', 'Cat_Schoon', 'Locatie', 'Prijs_Jouw', 'Prijs_CMS', 'Prijs_GoParts', 'Stock', 'KB']]
        df_onszelf.columns = ['Artikelnummer', 'Merk', 'Categorieën', 'Locatie', 'Jouw prijs', 'CMS prijs', 'GoParts prijs (19%)', 'Voorraad', 'Omschrijving']

        df_cms = df[['Title', 'Prijs_CMS', 'Stock']]
        df_cms.columns = ['Title', 'CMS Price', 'Stock']

        df_goparts = df[['Title', 'Merk', 'Cat_Schoon', 'Prijs_GoParts', 'Stock', 'KB']]
        df_goparts.columns = ['Artikelnummer', 'Merk', 'Categorieën', 'GoParts prijs in €', 'Voorraad', 'Omschrijving']

        # ----------------------------------------------------------
        # 9. Excel schrijven
        # ----------------------------------------------------------
        df_onszelf.to_excel(OUTPUT_ONSZELF, index=False)
        df_cms.to_excel(OUTPUT_CMS, index=False)
        df_goparts.to_excel(OUTPUT_GOPARTS, index=False)

        # ----------------------------------------------------------
        # 10. € formattering
        # ----------------------------------------------------------
        euro_style = NamedStyle(name='euro', number_format='€#,##0.00')

        def format_prijzen(file_path, cols):
            wb = load_workbook(file_path)
            ws = wb.active
            if 'euro' not in wb.named_styles:
                wb.add_named_style(euro_style)
            for col in cols:
                for row in range(2, ws.max_row + 1):
                    ws[f"{col}{row}"].style = 'euro'
            wb.save(file_path)

        format_prijzen(OUTPUT_ONSZELF, ['E', 'F', 'G'])
        format_prijzen(OUTPUT_CMS, ['B'])
        format_prijzen(OUTPUT_GOPARTS, ['D'])

        # ----------------------------------------------------------
        # 11. Log
        # ----------------------------------------------------------
        log_df = pd.DataFrame({
            'Log': [excel_safe(x) for x in log_lines],
            'Verwijderd': [excel_safe(x) for x in (verwijderd + [''] * (len(log_lines) - len(verwijderd)))]
        })

        log_df.to_excel(LOG_XLSX, index=False)


        return {
            'onszelf': OUTPUT_ONSZELF,
            'cms': OUTPUT_CMS,
            'goparts': OUTPUT_GOPARTS,
            'log': LOG_XLSX
        }
